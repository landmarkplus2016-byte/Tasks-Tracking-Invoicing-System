"""
Landmark Task & Invoicing Tracker — Backend Server
Run: python server.py --file "path/to/your/Total_Task_Tracking.xlsm"
Then open: http://localhost:5000
"""

import os, sys, json, argparse, threading, time
from datetime import datetime, date
from flask import Flask, jsonify, send_file, request
from flask_cors import CORS
import openpyxl

app = Flask(__name__)
CORS(app)

# ── Global cache ────────────────────────────────────────────────────────────
_cache = {"rows": [], "last_modified": None, "last_loaded": None, "row_count": 0}
_excel_path = None
_lock = threading.Lock()

# ── Field mapping: Excel column name → clean JSON key ───────────────────────
FIELD_MAP = {
    "ID#":                  "id",
    "Row":                  "row_num",
    "Job Code":             "job_code",
    "TX/RF":                "tx_rf",
    "Vendor":               "vendor",
    "Physical Site ID":     "physical_site_id",
    "Logical Site ID":      "logical_site_id",
    "Site Option":          "site_option",
    "Facing":               "facing",
    "Region":               "region",
    "Sub Region":           "sub_region",
    "Distance":             "distance",
    "Absolute Quantity":    "abs_qty",
    "Actual Quantity":      "actual_qty",
    "General Stream":       "stream",
    "Task Name":            "task_name",
    "Contractor":           "contractor",
    "Engineer's Name":      "engineer",
    "Line Item":            "line_item",
    "New Price":            "unit_price",
    "New Total Price":      "total_price",
    "Comments":             "comments",
    "Status":               "status",
    "Task Date":            "task_date",
    "VF Task owner":        "vf_owner",
    "PRQ":                  "prq",
    "Coordinator":          "coordinator",
    "Acceptance Status":    "acceptance_status",
    "FAC Date":             "fac_date",
    " Certificate #":       "cert_num",
    "Acceptance Week":      "acceptance_week",
    "TSR Sub#":             "tsr_sub",
    "PO status":            "po_status",
    "PO number":            "po_number",
    "VF Invoice #":         "vf_invoice",
    "1st Receiving Date":   "recv1_date",
    "2nd Receiving Date":   "recv2_date",
    "1st Receiving Amount": "recv1_amount",
    "1st Receiving Qnt":    "recv1_qty",
    "2nd Receiving Amount": "recv2_amount",
    "2nd Receiving Qnt":    "recv2_qty",
    "Remaining Amounts":    "remaining",
    "Submission Coments":   "submission_comments",
    "LMP":                  "lmp",
    "Contractor2":          "contractor_amount",
    "LMP Portion":          "lmp_portion",
    "Contractor Portion":   "contractor_portion",
    "Contractor Invoice #": "contractor_invoice",
    "Sent to Cost Control": "sent_cc",
    "Received from CC":     "recv_cc",
}

def fmt(v):
    """Serialize a cell value to a JSON-safe type."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float):
        if v != v:          # NaN
            return None
        if v == int(v):
            return int(v)
        return round(v, 4)
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v

def normalize_status(s):
    if not s:
        return s
    s = str(s).strip()
    mapping = {"done ": "Done", "assigned ": "Assigned",
               "cancelled": "Cancelled", "duplicated": "Cancelled",
               "dublicated ": "Cancelled"}
    return mapping.get(s.lower(), s)

def normalize_acceptance(s):
    if not s:
        return s
    s = str(s).strip().upper()
    if s in ("FAC", "TOC", "PAC"):
        return s
    return s

def load_excel(path, force=False):
    global _cache
    if not os.path.exists(path):
        print(f"[ERROR] File not found: {path}")
        return False
    try:
        mtime = os.path.getmtime(path)
        with _lock:
            if not force and _cache["last_modified"] == mtime and _cache["rows"]:
                return False   # no change
        
        print(f"[{datetime.now():%H:%M:%S}] Loading Excel…", end=" ", flush=True)
        wb = openpyxl.load_workbook(path, keep_vba=True, data_only=True)
        ws = wb["Invoicing Track"]

        # Row 4 = headers
        raw_headers = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
        headers = [h for h in raw_headers]

        rows = []
        for raw in ws.iter_rows(min_row=5, values_only=True):
            # Skip completely empty rows
            if not any(v for v in raw[:10] if v is not None):
                continue
            record = {}
            for col_name, val in zip(headers, raw):
                json_key = FIELD_MAP.get(col_name)
                if json_key:
                    record[json_key] = fmt(val)

            # Normalize status fields
            record["status"]            = normalize_status(record.get("status"))
            record["acceptance_status"] = normalize_acceptance(record.get("acceptance_status"))
            record["region"]            = str(record.get("region") or "").strip().title() or None

            # Compute remaining if formula returned None
            if record.get("remaining") is None:
                tp = record.get("total_price") or 0
                r1 = record.get("recv1_amount") or 0
                r2 = record.get("recv2_amount") or 0
                if isinstance(tp, (int, float)):
                    record["remaining"] = round(tp - r1 - r2, 2)

            rows.append(record)

        with _lock:
            _cache = {
                "rows": rows,
                "last_modified": mtime,
                "last_loaded": datetime.now().isoformat(),
                "row_count": len(rows),
            }
        print(f"{len(rows)} rows loaded ✓")
        return True
    except Exception as e:
        print(f"\n[ERROR] {e}")
        return False

def file_watcher():
    """Background thread: check for Excel changes every 15 s."""
    while True:
        time.sleep(15)
        if _excel_path:
            load_excel(_excel_path)

# ── Routes ──────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return send_file("index.html")

@app.route("/api/meta")
def meta():
    """Lightweight endpoint for change detection polling."""
    with _lock:
        return jsonify({
            "last_loaded": _cache["last_loaded"],
            "row_count":   _cache["row_count"],
        })

@app.route("/api/data")
def get_data():
    with _lock:
        return jsonify({
            "rows":        _cache["rows"],
            "row_count":   _cache["row_count"],
            "last_loaded": _cache["last_loaded"],
        })

@app.route("/api/reload")
def reload_data():
    load_excel(_excel_path, force=True)
    with _lock:
        return jsonify({"success": True, "row_count": _cache["row_count"],
                        "last_loaded": _cache["last_loaded"]})

# ── Entry point ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Landmark Task Tracker")
    parser.add_argument("--file", default="data.xlsm",
                        help="Path to your Excel tracking file (.xlsm)")
    parser.add_argument("--port", type=int, default=5000)
    args = parser.parse_args()

    _excel_path = args.file
    if not os.path.exists(_excel_path):
        print(f"[WARN] File '{_excel_path}' not found. Place your Excel file here or use --file.")
    else:
        load_excel(_excel_path, force=True)

    threading.Thread(target=file_watcher, daemon=True).start()
    print(f"\n🚀  Landmark Tracker running → http://localhost:{args.port}\n")
    app.run(port=args.port, debug=False, use_reloader=False)
